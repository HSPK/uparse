import os

import pandas as pd
from openpyxl import load_workbook

from uparse.schema import Chunk, Document

from ..pipeline import BaseTransform, Pipeline, State


class ExcelState(State):
    pass


class ParseExcel(BaseTransform[ExcelState]):
    def __init__(
        self, encoding: str | None = None, autodetect_encoding: bool = True, *args, **kwargs
    ):
        super().__init__(*args, **kwargs)
        self.encoding = encoding
        self.autodetect_encoding = autodetect_encoding

    async def transform(self, state, **kwargs):
        uri = state["uri"]
        chunks = []
        file_extension = os.path.splitext(uri)[-1].lower()
        if file_extension == ".xlsx":
            wb = load_workbook(uri, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = sheet.values
                cols = next(data)
                df = pd.DataFrame(data, columns=cols)

                df.dropna(how="all", inplace=True)

                for index, row in df.iterrows():
                    page_content = []
                    for col_index, (k, v) in enumerate(row.items()):
                        if pd.notna(v):
                            cell = sheet.cell(
                                row=index + 2, column=col_index + 1
                            )  # +2 to account for header and 1-based index
                            if cell.hyperlink:
                                value = f"[{v}]({cell.hyperlink.target})"
                                page_content.append(f'"{k}":"{value}"')
                            else:
                                page_content.append(f'"{k}":"{v}"')
                    chunks.append(
                        Chunk(
                            index=index,
                            content=";".join(page_content),
                            metadata={"source": uri},
                        )
                    )

        elif file_extension == ".xls":
            excel_file = pd.ExcelFile(uri, engine="xlrd")
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=sheet_name)
                df.dropna(how="all", inplace=True)

                for index, row in df.iterrows():
                    page_content = []
                    for k, v in row.items():
                        if pd.notna(v):
                            page_content.append(f'"{k}":"{v}"')
                    chunks.append(
                        Chunk(
                            index=index,
                            content=";".join(page_content),
                            metadata={"source": uri, "row": index},
                        )
                    )
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        doc = Document(metadata={"source": uri})
        doc.add_chunk(chunks)
        return doc


class ExcelPipeline(Pipeline):
    allowed_extensions = [".xls", ".xlsx"]
    """Load Excel files.


    Args:
        file_path: Path to the file to load.
    """

    def __init__(self, encoding=None, autodetect_encoding=True, *args, **kwargs):
        super().__init__(
            transforms=[ParseExcel(encoding=encoding, autodetect_encoding=autodetect_encoding)],
            *args,
            **kwargs,
        )
